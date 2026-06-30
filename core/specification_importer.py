from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping


NAME_HEADERS = {
    "название",
    "наименование",
    "имя",
    "деталь",
    "позиция",
    "обозначение",
}
QUANTITY_HEADERS = {
    "колво",
    "количество",
    "кол",
    "qty",
    "quantity",
}


@dataclass(frozen=True, slots=True)
class SpecificationItem:
    name: str
    quantity: int
    row: int


def load_quantity_specification(path: str | Path) -> dict[str, SpecificationItem]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Для чтения Excel-спецификаций нужна библиотека openpyxl. "
            "Обновите окружение или сборку программы."
        ) from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    items: dict[str, SpecificationItem] = {}
    for sheet in workbook.worksheets:
        header = _find_header_row(sheet)
        if header is None:
            continue
        header_row, name_column, quantity_column = header
        for row in range(header_row + 1, sheet.max_row + 1):
            raw_name = sheet.cell(row, name_column).value
            raw_quantity = sheet.cell(row, quantity_column).value
            name = _cell_text(raw_name)
            quantity = _parse_quantity(raw_quantity)
            if not name or quantity is None:
                continue
            key = normalize_spec_name(name)
            existing = items.get(key)
            if existing is None:
                items[key] = SpecificationItem(name=name, quantity=quantity, row=row)
            else:
                items[key] = SpecificationItem(
                    name=existing.name,
                    quantity=min(existing.quantity + quantity, 9999),
                    row=existing.row,
                )
    workbook.close()
    return items


def quantity_for_file(
    path: str | Path,
    items: Mapping[str, SpecificationItem],
) -> SpecificationItem | None:
    file_key = normalize_spec_name(Path(path).stem)
    exact = items.get(file_key)
    if exact is not None:
        return exact

    matches = [
        item
        for key, item in items.items()
        if key and (key in file_key or file_key in key)
    ]
    if len(matches) == 1:
        return matches[0]
    return None


def normalize_spec_name(value: str) -> str:
    text = str(value).casefold().replace("ё", "е")
    text = re.sub(r"\.[a-zа-я0-9]{1,5}$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(?:^|[\s_\-()[\]])(?:x|х|qty|q-ty|кол|kol|count)\s*[:=_-]?\s*\d{1,4}(?=$|[\s_\-().\[\]])", " ", text)
    text = re.sub(r"(?:^|[\s_\-()[\]])\d{1,4}\s*(?:шт|штук|pcs|pc|pieces)(?=$|[\s_\-().\[\]])", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _find_header_row(sheet: object) -> tuple[int, int, int] | None:
    max_scan_rows = min(getattr(sheet, "max_row", 0), 30)
    max_scan_columns = min(getattr(sheet, "max_column", 0), 30)
    for row in range(1, max_scan_rows + 1):
        name_column = 0
        quantity_column = 0
        for column in range(1, max_scan_columns + 1):
            header = _normalize_header(sheet.cell(row, column).value)
            if header in NAME_HEADERS:
                name_column = column
            if header in QUANTITY_HEADERS:
                quantity_column = column
        if name_column and quantity_column:
            return row, name_column, quantity_column
    return _guess_table_columns(sheet, max_scan_rows=max_scan_rows, max_scan_columns=max_scan_columns)


def _guess_table_columns(
    sheet: object,
    *,
    max_scan_rows: int,
    max_scan_columns: int,
) -> tuple[int, int, int] | None:
    best: tuple[int, int, int, int] | None = None
    max_row = getattr(sheet, "max_row", 0)
    for start_row in range(1, max_scan_rows + 1):
        rows_to_check = range(start_row, min(max_row, start_row + 15) + 1)
        for name_column in range(1, max_scan_columns + 1):
            name_score = sum(
                1
                for row in rows_to_check
                if _looks_like_part_name(sheet.cell(row, name_column).value)
            )
            if name_score < 2:
                continue
            for quantity_column in range(1, max_scan_columns + 1):
                if quantity_column == name_column:
                    continue
                quantity_score = sum(
                    1
                    for row in rows_to_check
                    if _parse_quantity(sheet.cell(row, quantity_column).value) is not None
                )
                score = min(name_score, quantity_score)
                if score < 2:
                    continue
                if best is None or score > best[0]:
                    best = (score, start_row - 1, name_column, quantity_column)
    if best is None:
        return None
    _score, header_row, name_column, quantity_column = best
    return header_row, name_column, quantity_column


def _looks_like_part_name(value: object) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    if _parse_quantity(text) is not None and len(text) <= 6:
        return False
    return bool(re.search(r"[A-Za-zА-Яа-я]", text)) and len(normalize_spec_name(text)) >= 3


def _normalize_header(value: object) -> str:
    text = _cell_text(value).casefold().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_quantity(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        quantity = int(value)
        return quantity if quantity >= 1 else None
    match = re.search(r"\d{1,4}", str(value))
    if match is None:
        return None
    quantity = int(match.group(0))
    return quantity if quantity >= 1 else None
